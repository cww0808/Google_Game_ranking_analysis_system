from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Malgun Gothic", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Malgun Gothic", size=10)
RISE_FILL = PatternFill("solid", fgColor="E2F0D9")
FALL_FILL = PatternFill("solid", fgColor="FCE4D6")


def style_sheet(sheet, widths: dict[str, float], table_name: str) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top")
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    if sheet.max_row > 1:
        table = Table(displayName=table_name, ref=f"A1:{sheet.cell(1, sheet.max_column).column_letter}{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def get_previous_snapshot(connection: sqlite3.Connection, snapshot: str) -> str | None:
    row = connection.execute(
        """
        SELECT MAX(collected_date) AS previous_snapshot
        FROM game_snapshots
        WHERE collected_date < ?
        """,
        (snapshot,),
    ).fetchone()
    return row["previous_snapshot"] if row and row["previous_snapshot"] else None


def change_status(current_rank: int | None, previous_rank: int | None, exists_current: bool = True) -> str:
    if not exists_current:
        return "이탈"
    if previous_rank is None:
        return "신규"
    if current_rank is None:
        return "이탈"
    if previous_rank > current_rank:
        return "상승"
    if previous_rank < current_rank:
        return "하락"
    return "유지"


def rank_change(current_rank: int | None, previous_rank: int | None) -> int | None:
    if current_rank is None or previous_rank is None:
        return None
    return previous_rank - current_rank


def write_rank_change_sheet(
    workbook: Workbook,
    connection: sqlite3.Connection,
    snapshot: str,
    previous_snapshot: str | None,
) -> None:
    sheet = workbook.create_sheet("순위 변동 요약", 0)
    sheet.append(["수집 회차", snapshot])
    sheet.append(["비교 기준", previous_snapshot or "없음 - 첫 수집"])
    sheet.append([])
    sheet.append(["구분", "현재 순위", "이전 순위", "순위 변화", "개발사", "게임명", "패키지명", "장르"])

    if previous_snapshot is None:
        current_rows = connection.execute(
            """
            SELECT rank, developer, title, app_id, genre
            FROM game_snapshots
            WHERE collected_date = ?
            ORDER BY rank
            """,
            (snapshot,),
        )
        for row in current_rows:
            sheet.append(["신규", row["rank"], None, None, row["developer"], row["title"], row["app_id"], row["genre"]])
    else:
        rows = connection.execute(
            """
            WITH current AS (
                SELECT app_id, rank, developer, title, genre
                FROM game_snapshots
                WHERE collected_date = ?
            ),
            previous AS (
                SELECT app_id, rank, developer, title, genre
                FROM game_snapshots
                WHERE collected_date = ?
            )
            SELECT
                c.rank AS current_rank,
                p.rank AS previous_rank,
                COALESCE(c.developer, p.developer) AS developer,
                COALESCE(c.title, p.title) AS title,
                COALESCE(c.app_id, p.app_id) AS app_id,
                COALESCE(c.genre, p.genre) AS genre,
                CASE WHEN c.app_id IS NULL THEN 0 ELSE 1 END AS exists_current
            FROM current c
            LEFT JOIN previous p ON p.app_id = c.app_id
            UNION ALL
            SELECT
                NULL AS current_rank,
                p.rank AS previous_rank,
                p.developer,
                p.title,
                p.app_id,
                p.genre,
                0 AS exists_current
            FROM previous p
            LEFT JOIN current c ON c.app_id = p.app_id
            WHERE c.app_id IS NULL
            """,
            (snapshot, previous_snapshot),
        )
        decorated = []
        for row in rows:
            current_rank = row["current_rank"]
            previous_rank = row["previous_rank"]
            exists_current = bool(row["exists_current"])
            status = change_status(current_rank, previous_rank, exists_current)
            decorated.append(
                [
                    status,
                    current_rank,
                    previous_rank,
                    rank_change(current_rank, previous_rank),
                    row["developer"],
                    row["title"],
                    row["app_id"],
                    row["genre"],
                ]
            )
        status_order = {"상승": 0, "하락": 1, "신규": 2, "이탈": 3, "유지": 4}
        decorated.sort(
            key=lambda row: (
                status_order.get(row[0], 9),
                -(row[3] or -999) if row[0] == "상승" else (row[3] or 999),
                row[1] or 999,
                row[2] or 999,
            )
        )
        for row in decorated:
            sheet.append(row)

    for row_idx in (1, 2):
        for cell in sheet[row_idx]:
            cell.font = Font(name="Malgun Gothic", bold=True)
    for cell in sheet[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top")
    widths = {"A": 10, "B": 11, "C": 11, "D": 11, "E": 24, "F": 32, "G": 34, "H": 18}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    if sheet.max_row >= 5:
        table = Table(displayName="RankChangesTable", ref=f"A4:H{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
        sheet.conditional_formatting.add(f"D5:D{sheet.max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=RISE_FILL))
        sheet.conditional_formatting.add(f"D5:D{sheet.max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=FALL_FILL))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", default="data/google_play_games.db")
    parser.add_argument("--snapshot", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    connection = sqlite3.connect(args.db)
    connection.row_factory = sqlite3.Row
    previous_snapshot = get_previous_snapshot(connection, args.snapshot)

    workbook = Workbook()
    reviews_sheet = workbook.active
    reviews_sheet.title = "리뷰"
    users_sheet = workbook.create_sheet("사용자 관계")
    games_sheet = workbook.create_sheet("게임 요약")

    write_rank_change_sheet(workbook, connection, args.snapshot, previous_snapshot)

    reviews_sheet.append(
        [
            "개발사", "게임명", "패키지명", "수집 당시 순위", "사용자 식별키",
            "닉네임", "별점", "평가글", "작성일", "추천 수", "앱 버전",
            "리뷰 ID", "프로필 이미지 URL", "식별 방식", "리뷰 URL",
        ]
    )
    review_rows = connection.execute(
        """
        SELECT
            gs.developer, gr.game_title, gr.app_id, gs.rank, gr.reviewer_key,
            gr.user_name_at_review, gr.score, gr.review_text, gr.review_date,
            gr.thumbs_up, gr.app_version, gr.review_id, gr.user_image_at_review,
            r.identity_method, gr.review_url
        FROM game_reviews gr
        JOIN reviewers r ON r.reviewer_key = gr.reviewer_key
        LEFT JOIN game_snapshots gs
          ON gs.app_id = gr.app_id AND gs.collected_date = ?
        ORDER BY COALESCE(gs.rank, 999), gr.review_date DESC
        """,
        (args.snapshot,),
    )
    for row in review_rows:
        reviews_sheet.append(list(row))

    users_sheet.append(
        [
            "사용자 식별키", "최근 닉네임", "식별 방식", "리뷰 수",
            "평가한 게임 수", "최초 리뷰일", "최근 리뷰일", "평가 게임 목록",
        ]
    )
    user_rows = connection.execute(
        """
        SELECT
            r.reviewer_key, r.latest_user_name, r.identity_method,
            COUNT(gr.review_id), COUNT(DISTINCT gr.app_id),
            MIN(gr.review_date), MAX(gr.review_date),
            GROUP_CONCAT(DISTINCT gr.game_title)
        FROM reviewers r
        JOIN game_reviews gr ON gr.reviewer_key = r.reviewer_key
        GROUP BY r.reviewer_key
        ORDER BY COUNT(gr.review_id) DESC, COUNT(DISTINCT gr.app_id) DESC
        """
    )
    for row in user_rows:
        users_sheet.append(list(row))

    games_sheet.append(
        [
            "순위", "이전 순위", "순위 변화", "변동 상태", "개발사", "게임명", "패키지명", "수집 리뷰 수",
            "고유 사용자 수", "평균 별점", "1점 리뷰", "5점 리뷰",
        ]
    )
    game_rows = connection.execute(
        """
        SELECT
            gs.rank, prev.rank AS previous_rank, gs.developer, gs.title, gs.app_id,
            COUNT(gr.review_id) AS review_count, COUNT(DISTINCT gr.reviewer_key) AS reviewer_count,
            ROUND(AVG(gr.score), 2) AS avg_score,
            SUM(CASE WHEN gr.score = 1 THEN 1 ELSE 0 END) AS one_star_reviews,
            SUM(CASE WHEN gr.score = 5 THEN 1 ELSE 0 END) AS five_star_reviews
        FROM game_snapshots gs
        LEFT JOIN game_snapshots prev
          ON prev.app_id = gs.app_id AND prev.collected_date = ?
        LEFT JOIN game_reviews gr ON gr.app_id = gs.app_id
        WHERE gs.collected_date = ?
        GROUP BY gs.app_id, gs.rank, prev.rank, gs.developer, gs.title
        ORDER BY gs.rank
        """,
        (previous_snapshot, args.snapshot),
    )
    for row in game_rows:
        current_rank = row["rank"]
        previous_rank = row["previous_rank"]
        games_sheet.append(
            [
                current_rank,
                previous_rank,
                rank_change(current_rank, previous_rank),
                change_status(current_rank, previous_rank),
                row["developer"],
                row["title"],
                row["app_id"],
                row["review_count"],
                row["reviewer_count"],
                row["avg_score"],
                row["one_star_reviews"],
                row["five_star_reviews"],
            ]
        )

    style_sheet(
        reviews_sheet,
        {"A": 23, "B": 28, "C": 30, "D": 12, "E": 36, "F": 20, "G": 8, "H": 58, "I": 21, "J": 10, "K": 12, "L": 38, "M": 48, "N": 25, "O": 48},
        "ReviewsTable",
    )
    reviews_sheet.column_dimensions["H"].width = 58
    for cell in reviews_sheet["H"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    style_sheet(
        users_sheet,
        {"A": 36, "B": 20, "C": 25, "D": 10, "E": 14, "F": 21, "G": 21, "H": 65},
        "ReviewersTable",
    )
    users_sheet.conditional_formatting.add(
        f"D2:D{users_sheet.max_row}",
        DataBarRule(start_type="min", end_type="max", color="5B9BD5"),
    )

    style_sheet(
        games_sheet,
        {"A": 9, "B": 11, "C": 11, "D": 11, "E": 24, "F": 30, "G": 32, "H": 14, "I": 15, "J": 12, "K": 11, "L": 11},
        "GamesTable",
    )
    games_sheet.conditional_formatting.add(
        f"C2:C{games_sheet.max_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )
    games_sheet.conditional_formatting.add(
        f"J2:J{games_sheet.max_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
